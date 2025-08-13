#!/usr/bin/env python
# coding: utf-8

# In[29]:


###############################     Lancement des tests   ###############################

# Le programme fonctionne en 2 étapes :
# 1) Il va lire la table et va créer des variables qui vont prendre la valeur 1 si le test est OK, 0 sinon.
# 2) Ensuite, il va sommer chaque colonne et la table finale va  contenir le nombre d'enregistrements OK pour un test donné.

# Tables :
# - PERIM : Contient le pérmètre permettant l'analyse
# - TEST : Table en sortie de l'étape 1
# - RESULTS : Table en sortie de l'étape 2 et qui permet de sortir les chiffres

# ***************************************************/


# /********************************************/
# /*****###############* DTM MidCorp : Table PTF POLICY #########################################******/
# /********************************************/

import pandas as pd

get_ipython().run_line_magic('run', 'LIEN_FONCT.ipynb')


# In[31]:


##################################### POLICY  ###################################################


PERIM_PTF_MC = pd.DataFrame()
PERIM_PTF_MC = PTF_MC.copy()

#PERIM_PTF_MC

varPTFPOL = ["NOPOL", "NOCLT", "CDSITP", "DTEFSITT", "DTEXPIR", "DTRESILP", "DTCREPOL", "LCI_100", "LCI_100_IND", "SMP_100", "SMP_100_IND", "PRIMES_PTF", "PCRABCOM", "POSRISQ", "CMARCH", "CSEG", "CSSSEG", "CDNATP"]

#######################   ETAPE 1 ######################################

TEST_MIDCORP_POLICY_PTF = pd.DataFrame()
TEST_MIDCORP_POLICY_PTF = PERIM_PTF_MC[varPTFPOL].copy()

TEST_MIDCORP_POLICY_PTF['COUNT_LIGNES'] = 1

##########   NOPOL #####################

COMPLETENESS(TEST_MIDCORP_POLICY_PTF, "NOPOL")
TECH_VAL_FORMAT_VAR(TEST_MIDCORP_POLICY_PTF, "NOPOL")
TECH_VALIDITY_LENGTH(TEST_MIDCORP_POLICY_PTF, "NOPOL")

#############################   NOCLT  #############################

COMPLETENESS(TEST_MIDCORP_POLICY_PTF, "NOCLT")
TECH_VAL_FORMAT_VAR(TEST_MIDCORP_POLICY_PTF, "NOCLT")
TECH_VALIDITY_LENGTH(TEST_MIDCORP_POLICY_PTF, "NOCLT")

##################################   CDSITP   ###############################

COMPLETENESS(TEST_MIDCORP_POLICY_PTF, "CDSITP")
BUSINESS_VALIDY (TEST_MIDCORP_POLICY_PTF, "CDSITP")

#########################   DTEFSITT   #######################

COMPLETENESS(TEST_MIDCORP_POLICY_PTF, "DTEFSITT")
BUSINESS_VALIDY (TEST_MIDCORP_POLICY_PTF, "DTEFSITT")


########################################   CDNATP ( Conrespond aux nbre de ligne en entré pour la variable DTEXPIR)   ########################################
TEST_MIDCORP_POLICY_PTF["ENTRE_DTEXPIR"] = np.where(TEST_MIDCORP_POLICY_PTF["CDNATP"] == 'T', 1, 0)

#########################   DTEXPIR    ###########################

COMPLETENESS(TEST_MIDCORP_POLICY_PTF, "DTEXPIR")
BUSINESS_VALIDY (TEST_MIDCORP_POLICY_PTF, "DTEXPIR")

#########################    DTRESILP   ##################################

#Spécifique car on souhaite regarder uniquemment les enregistrement qui ont un statut de police à Cancelled / RESILIEE 

#COUNT_POLICY_CANCELLED
TEST_MIDCORP_POLICY_PTF["COUNT_POLICY_CANCELLED"] = np.where(TEST_MIDCORP_POLICY_PTF["CDSITP"] == '3', 1, 0)

TEST_MIDCORP_POLICY_PTF["COMP_DTRESILP"] = 0
TEST_MIDCORP_POLICY_PTF.loc[TEST_MIDCORP_POLICY_PTF["COUNT_POLICY_CANCELLED"] == 1, "COMP_DTRESILP"] = TEST_MIDCORP_POLICY_PTF.loc[TEST_MIDCORP_POLICY_PTF["COUNT_POLICY_CANCELLED"] == 1, "DTRESILP"].apply(lambda t: int(not (pd.isna(t) | (t == ''))))

#BUS_VAL_DTRESILP
TEST_MIDCORP_POLICY_PTF["BUS_VAL_DTRESILP"] = 0

TEST_MIDCORP_POLICY_PTF["A"] = np.where(((TEST_MIDCORP_POLICY_PTF["COMP_DTRESILP"] == 1) & (TEST_MIDCORP_POLICY_PTF["COMP_DTEFSITT"] == 1)), 1, 0)
df = TEST_MIDCORP_POLICY_PTF[TEST_MIDCORP_POLICY_PTF["A"] == 1]
df.loc[df["COUNT_POLICY_CANCELLED"] == 1, "BUS_VAL_DTRESILP"] = df[df["COUNT_POLICY_CANCELLED"] == 1].apply(lambda t: pd.to_datetime(t["DTRESILP"]) >= pd.to_datetime(t["DTEFSITT"]), axis = 1)
TEST_MIDCORP_POLICY_PTF["BUS_VAL_DTRESILP"]  =  df["BUS_VAL_DTRESILP"].astype(int)  # je le transforme en int ici d'abord comme ca il ne reste que les NaN à changer

TEST_MIDCORP_POLICY_PTF["BUS_VAL_DTRESILP"].fillna(0, inplace=True) # je le transforme ici si non il va me transforme aussi les Fals en 0
TEST_MIDCORP_POLICY_PTF["BUS_VAL_DTRESILP"] = TEST_MIDCORP_POLICY_PTF["BUS_VAL_DTRESILP"].astype(int)

################## DTCREPOL #####################

COMPLETENESS(TEST_MIDCORP_POLICY_PTF, "DTCREPOL")
BUSINESS_VALIDY (TEST_MIDCORP_POLICY_PTF, "DTCREPOL")


TEST_MIDCORP_POLICY_PTF["COUNT_POLICY_ENCOURS"] = np.where(TEST_MIDCORP_POLICY_PTF["CDSITP"] == "1", 1, 0)

###############################    PCRABCOM   ############################3

COMPLETENESS(TEST_MIDCORP_POLICY_PTF, "PCRABCOM")
BUSINESS_VALIDY (TEST_MIDCORP_POLICY_PTF, "PCRABCOM")

############################ POSRISQ   ##############################

COMPLETENESS(TEST_MIDCORP_POLICY_PTF, "POSRISQ")

#################### LCI_100_IND ########################################

COMPLETENESS(TEST_MIDCORP_POLICY_PTF, "LCI_100_IND")	
BUSINESS_VALIDY (TEST_MIDCORP_POLICY_PTF, "LCI_100_IND")

###############   LCI_100 #################################

COMPLETENESS(TEST_MIDCORP_POLICY_PTF, "LCI_100")
BUSINESS_VALIDY (TEST_MIDCORP_POLICY_PTF, "LCI_100")

#################### SMP_100_IND ##################################

COMPLETENESS(TEST_MIDCORP_POLICY_PTF, "SMP_100_IND")
BUSINESS_VALIDY (TEST_MIDCORP_POLICY_PTF, "SMP_100_IND")

######################   SMP_100 ###################################

COMPLETENESS(TEST_MIDCORP_POLICY_PTF, "SMP_100")
BUSINESS_VALIDY (TEST_MIDCORP_POLICY_PTF, "SMP_100")

################  PRIMES_PTF ################

##Spécifique car il y a 3 BUSINESS TEST */

COMPLETENESS(TEST_MIDCORP_POLICY_PTF, "PRIMES_PTF")

TEST_MIDCORP_POLICY_PTF["BUS_VAL_PRIMES_PTF_0"] = 0
TEST_MIDCORP_POLICY_PTF.loc[TEST_MIDCORP_POLICY_PTF["COMP_PRIMES_PTF"] == 1, "BUS_VAL_PRIMES_PTF_0"] = np.where(TEST_MIDCORP_POLICY_PTF.loc[TEST_MIDCORP_POLICY_PTF["COMP_PRIMES_PTF"] == 1, "PRIMES_PTF"] > 0, 1, 0)


TEST_MIDCORP_POLICY_PTF["BUS_VAL_PRIMES_PTF_MAX"] = 0
TEST_MIDCORP_POLICY_PTF.loc[TEST_MIDCORP_POLICY_PTF["COMP_PRIMES_PTF"] == 1, "BUS_VAL_PRIMES_PTF_MAX"] = np.where(TEST_MIDCORP_POLICY_PTF.loc[TEST_MIDCORP_POLICY_PTF["COMP_PRIMES_PTF"] == 1, "PRIMES_PTF"] <= 1000000, 1, 0)


TEST_MIDCORP_POLICY_PTF["BUS_VAL_PRIMES_PTF_MIN"] = 0
TEST_MIDCORP_POLICY_PTF.loc[TEST_MIDCORP_POLICY_PTF["COMP_PRIMES_PTF"] == 1, "BUS_VAL_PRIMES_PTF_MIN"] = np.where(TEST_MIDCORP_POLICY_PTF.loc[TEST_MIDCORP_POLICY_PTF["COMP_PRIMES_PTF"] == 1, "PRIMES_PTF"] > 1000, 1, 0)

########################################   (CMARCH, CSEG, CSSSEG) = CCC   ########################################

COMPLETENESS(TEST_MIDCORP_POLICY_PTF, "CMARCH")
COMPLETENESS(TEST_MIDCORP_POLICY_PTF, "CSEG")
COMPLETENESS(TEST_MIDCORP_POLICY_PTF, "CSSSEG")

TEST_MIDCORP_POLICY_PTF["COMP_CCC"] = np.where((TEST_MIDCORP_POLICY_PTF["COMP_CMARCH"] == 1)  & (TEST_MIDCORP_POLICY_PTF["COMP_CSEG"] == 1
) & (TEST_MIDCORP_POLICY_PTF["COMP_CSSSEG"] == 1
), 1, 0)

BUSINESS_VALIDY(TEST_MIDCORP_POLICY_PTF, "CMARCH")
BUSINESS_VALIDY(TEST_MIDCORP_POLICY_PTF, "CSEG")
BUSINESS_VALIDY(TEST_MIDCORP_POLICY_PTF, "CSSSEG")

TEST_MIDCORP_POLICY_PTF["BUS_VAL_CCC"] = np.where((TEST_MIDCORP_POLICY_PTF["BUS_VAL_CMARCH"] == 1)  & (TEST_MIDCORP_POLICY_PTF["BUS_VAL_CSEG"] == 1
) & (TEST_MIDCORP_POLICY_PTF["BUS_VAL_CSSSEG"] == 1
), 1, 0)


# In[32]:


### ICI JE SUPPRIME TOUT ET JE SORS LA TABLE FINALE DE ETAPE 1 

TEST_MIDCORP_POLICY_PTF = TEST_MIDCORP_POLICY_PTF.drop(columns = varPTFPOL) 


# In[40]:


#TEST_MIDCORP_POLICY_PTF


# In[34]:


######################################################   ETAPE 2  ################################

RESULTS_TT_MIDCORP_POLICY_PTF = pd.DataFrame()


############## Nombre de ligne au total ###########################
RESULTS_TT_MIDCORP_POLICY_PTF = pd.DataFrame({"SUM_COUNT_LIGNES" : [(TEST_MIDCORP_POLICY_PTF["COUNT_LIGNES"]).sum()]})


###############  NOPOL   ####################################
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["NOPOL"]}_COMP'] = TEST_MIDCORP_POLICY_PTF["COMP_NOPOL"].sum()
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["NOPOL"]}_FORMAT'] = TEST_MIDCORP_POLICY_PTF["FORMAT_NOPOL"].sum()
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["NOPOL"]}_LENGTH'] = TEST_MIDCORP_POLICY_PTF["LENGTH_NOPOL"].sum()


#############    NOCLT    ############################
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["NOCLT"]}_COMP'] = TEST_MIDCORP_POLICY_PTF["COMP_NOCLT"].sum()
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["NOCLT"]}_FORMAT'] = TEST_MIDCORP_POLICY_PTF["FORMAT_NOCLT"].sum()
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["NOCLT"]}_LENGTH'] = TEST_MIDCORP_POLICY_PTF["LENGTH_NOCLT"].sum()


#############    CDSITP   ##############################
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["CDSITP"]}_COMP'] = TEST_MIDCORP_POLICY_PTF["COMP_CDSITP"].sum()
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["CDSITP"]}_BUS_VAL'] = TEST_MIDCORP_POLICY_PTF["BUS_VAL_CDSITP"].sum()


#############    DTEFSITT  ###############################
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["DTEFSITT"]}_COMP'] = TEST_MIDCORP_POLICY_PTF["COMP_DTEFSITT"].sum()
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["DTEFSITT"]}_BUS_VAL'] = TEST_MIDCORP_POLICY_PTF["BUS_VAL_DTEFSITT"].sum()

########################################   CDNATP ( Conrespond aux nbre de ligne en entré pour la variable DTEXPIR)   ########################################
RESULTS_TT_MIDCORP_POLICY_PTF["SUM_ENTRE_DTEXPIR"] = TEST_MIDCORP_POLICY_PTF["ENTRE_DTEXPIR"].sum()


############    DTEXPIR    #################################
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["DTEXPIR"]}_COMP'] = TEST_MIDCORP_POLICY_PTF["COMP_DTEXPIR"].sum()
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["DTEXPIR"]}_BUS_VAL'] = TEST_MIDCORP_POLICY_PTF["BUS_VAL_DTEXPIR"].sum()


###########    DTRESILP    ##################################
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["COUNT_POLICY_CANCELLED"]}_COMP'] = TEST_MIDCORP_POLICY_PTF["COUNT_POLICY_CANCELLED"].sum()
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["DTRESILP"]}_COMP'] = TEST_MIDCORP_POLICY_PTF["COMP_DTRESILP"].sum()
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["DTRESILP"]}_BUS_VAL'] = TEST_MIDCORP_POLICY_PTF["BUS_VAL_DTRESILP"].sum()	


##################     DTCREPOL   ################################
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["DTCREPOL"]}_COMP'] = TEST_MIDCORP_POLICY_PTF["COMP_DTCREPOL"].sum()	
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["DTCREPOL"]}_BUS_VAL'] = TEST_MIDCORP_POLICY_PTF["BUS_VAL_DTCREPOL"].sum()


RESULTS_TT_MIDCORP_POLICY_PTF["SUM_COUNT_POLICY_ENCOURS"] = TEST_MIDCORP_POLICY_PTF["COUNT_POLICY_ENCOURS"].sum()	


###################       LCI_100_IND      ####################################
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["LCI_100_IND"]}_COMP'] = TEST_MIDCORP_POLICY_PTF["COMP_LCI_100_IND"].sum()	
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["LCI_100_IND"]}_BUS_VAL'] = TEST_MIDCORP_POLICY_PTF["BUS_VAL_LCI_100_IND"].sum()	


##################     LCI_100      ##################################
RESULTS_TT_MIDCORP_POLICY_PTF["SUM_COMP_LCI_100"] = TEST_MIDCORP_POLICY_PTF["COMP_LCI_100"].sum()	
RESULTS_TT_MIDCORP_POLICY_PTF["SUM_BUS_VAL_LCI_100"] = TEST_MIDCORP_POLICY_PTF["BUS_VAL_LCI_100"].sum()	


#################     SMP_100_IND     ############################
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["SMP_100_IND"]}_COMP'] = TEST_MIDCORP_POLICY_PTF["COMP_SMP_100_IND"].sum()	
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["SMP_100_IND"]}_BUS_VAL'] = TEST_MIDCORP_POLICY_PTF["BUS_VAL_SMP_100_IND"].sum()	


###############   SMP_100     #########################
RESULTS_TT_MIDCORP_POLICY_PTF["SUM_COMP_SMP_100"] = TEST_MIDCORP_POLICY_PTF["COMP_SMP_100"].sum()	
RESULTS_TT_MIDCORP_POLICY_PTF["SUM_BUS_VAL_SMP_100"] = TEST_MIDCORP_POLICY_PTF["BUS_VAL_SMP_100"].sum()	


################     PRIMES_PTF   #########################
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["PRIMES_PTF"]}_COMP'] = TEST_MIDCORP_POLICY_PTF["COMP_PRIMES_PTF"].sum()	
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["PRIMES_PTF"]}_BUS_VAL_0'] = TEST_MIDCORP_POLICY_PTF["BUS_VAL_PRIMES_PTF_0"].sum()	
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["PRIMES_PTF"]}_BUS_VAL_MAX'] = TEST_MIDCORP_POLICY_PTF["BUS_VAL_PRIMES_PTF_MAX"].sum()	
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["PRIMES_PTF"]}_BUS_VAL_MIN'] = TEST_MIDCORP_POLICY_PTF["BUS_VAL_PRIMES_PTF_MIN"].sum()	


#################     PCRABCOM    ###########################
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["PCRABCOM"]}_COMP'] = TEST_MIDCORP_POLICY_PTF["COMP_PCRABCOM"].sum()	
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["PCRABCOM"]}_BUS_VAL'] = TEST_MIDCORP_POLICY_PTF["BUS_VAL_PCRABCOM"].sum()


#################      POSRISQ     ####################
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["POSRISQ"]}_COMP'] = TEST_MIDCORP_POLICY_PTF["COMP_POSRISQ"].sum()	

# BUS_VAL POSRISQ spécifique
Damo = PTF_MC[PTF_MC["POSRISQ"].isin(LISTE_CODE_POSTAUX["CODE_POSTAL"])]
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["POSRISQ"]}_BUS_VAL'] = len(Damo)

####################### (CMARCH, CSEG, CSSSEG) = CCC ###########################
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["CCC"]}_COMP'] = TEST_MIDCORP_POLICY_PTF["COMP_CCC"].sum()
RESULTS_TT_MIDCORP_POLICY_PTF[f'{REFMIDCORP["CCC"]}_BUS_VAL'] = TEST_MIDCORP_POLICY_PTF["BUS_VAL_CCC"].sum()


# In[41]:


#RESULTS_TT_MIDCORP_POLICY_PTF


# In[37]:


from openpyxl import load_workbook
import pandas as pd

chemin = './test/DataFitness2024_DATAFY_P4D_Didier 1 (version 1).xlsx'

workbook = load_workbook(chemin)


NouvelleFeuille = f"MIDCORP_POLICY_PTF_{annee}{mois}"


nouvelle_feuille = workbook.create_sheet(title= NouvelleFeuille)

nouvelle_feuille.append(list(RESULTS_TT_MIDCORP_POLICY_PTF.columns))

for index, row in RESULTS_TT_MIDCORP_POLICY_PTF.iterrows():
    nouvelle_feuille.append(list(row))

workbook.save(chemin)

print(f"Le DataFrame a été ajouté dans {chemin} sous la feuille '{NouvelleFeuille}'")


workbook.close()


# In[38]:


################   Spécifique pour tester le business validity des codes postaux  ########################################

Damo = PTF_MC[PTF_MC["POSRISQ"].isin(LISTE_CODE_POSTAUX["CODE_POSTAL"])]

RESULTS_BUS_VAL_POSTISQ = pd.DataFrame({f'{REFMIDCORP["POSRISQ"]}_BUS_VAL' : [len(Damo)]})


# In[39]:


RESULTS_BUS_VAL_POSTISQ   #cette DF a été rajouté sur la plus grande DF nommé "RESULTS_TT_MIDCORP_POLICY_PTF"


# In[12]:


######################################################################################################################
############################     DTM MidCorp : Table MVT POLICY      ######################################
#######################################################################################################################


##############################    PERIMETRE     ##################################################################


PERIM_MIDCORP_MVT = pd.DataFrame()

Didier = pd.merge(MVT_MC, PTF_MC, on='NOPOL', how='inner')
PERIM_MIDCORP_MVT = Didier[['DTTREVEN', 'DTEFSITT', 'DTEXPIR']]

#PERIM_MIDCORP_MVT


####################################   ETAPE 1    ##################################################

TEST_MIDCORP_POLICY_MVT = PERIM_MIDCORP_MVT[['DTTREVEN', 'DTEFSITT', 'DTEXPIR']].copy()
TEST_MIDCORP_POLICY_MVT["COUNT_LIGNE"] = 1

COMPLETENESS(TEST_MIDCORP_POLICY_MVT, "DTTREVEN")
BUSINESS_VALIDY(TEST_MIDCORP_POLICY_MVT, "DTTREVEN")

TEST_MIDCORP_POLICY_MVT = TEST_MIDCORP_POLICY_MVT.drop(columns = ['DTTREVEN', 'DTEFSITT', 'DTEXPIR'])

#TEST_MIDCORP_POLICY_MVT


############################# ETAPE 2 ###################################

RESULTS_MIDCORP_POLICY_MVT = pd.DataFrame({"SUM_COUNT_LIGNE" : [TEST_MIDCORP_POLICY_MVT["COUNT_LIGNE"].sum()]})
RESULTS_MIDCORP_POLICY_MVT[f'{REFMIDCORP["DTTREVEN"]}_COMP'] = TEST_MIDCORP_POLICY_MVT["COMP_DTTREVEN"].sum()
RESULTS_MIDCORP_POLICY_MVT[f'{REFMIDCORP["DTTREVEN"]}_BUS_VAL'] = TEST_MIDCORP_POLICY_MVT["BUS_VAL_DTTREVEN"].sum()  


# In[13]:


#RESULTS_MIDCORP_POLICY_MVT


# In[15]:


from openpyxl import load_workbook
import pandas as pd

chemin = './test/DataFitness2024_DATAFY_P4D_Didier 1 (version 1).xlsx'

workbook = load_workbook(chemin)


NouvelleFeuille = f"MIDCORP_POLICY_MVT_{annee}{mois}"


nouvelle_feuille = workbook.create_sheet(title= NouvelleFeuille)

nouvelle_feuille.append(list(RESULTS_MIDCORP_POLICY_MVT.columns))

for index, row in RESULTS_MIDCORP_POLICY_MVT.iterrows():
    nouvelle_feuille.append(list(row))

workbook.save(chemin)

print(f"Le DataFrame a été ajouté dans {chemin} sous la feuille '{NouvelleFeuille}'")


workbook.close()


# In[17]:


##########################################       IMS    ####################################################################
####################################################################################################################

####################################### PERIMETRE ####################################################################
#Concaténation de l'ensemble des tables du périmètre dans une seule table */


PERIM_PTF_IMS = pd.DataFrame()


#concat = pd.concat()
filtered_PTF14 = PTF14.loc[FILTRE_PTF(PTF14), ['NOPOL'] + VAR_PTF_GAR]
filtered_PTF34 = PTF34.loc[FILTRE_PTF(PTF34), ['NOPOL'] + VAR_PTF_GAR]
filtered_PTF15 = PTF15.loc[FILTRE_PTF(PTF15), ['NOPOL'] + VAR_PTF_GAR]
filtered_PTF35 = PTF35.loc[FILTRE_PTF(PTF35), ['NOPOL'] + VAR_PTF_GAR]
filtered_PTF1M41 = PTF1M41.loc[FILTRE_PTF(PTF1M41), ['NOPOL'] + VAR_PTF_GAR]
filtered_PTF3M41 = PTF3M41.loc[FILTRE_PTF(PTF3M41), ['NOPOL'] + VAR_PTF_GAR]
filtered_PTF1MPI = PTF1MPI.loc[FILTRE_PTF(PTF1MPI), ['NOPOL'] + VAR_PTF_GAR]
filtered_PTF3MPI = PTF3MPI.loc[FILTRE_PTF(PTF3MPI), ['NOPOL'] + VAR_PTF_GAR]

PERIM_PTF_IMS = pd.concat([filtered_PTF14, filtered_PTF34, filtered_PTF15, filtered_PTF35, filtered_PTF1M41, filtered_PTF3M41, filtered_PTF1MPI, filtered_PTF3MPI], ignore_index=True)  
#PERIM_PTF_IMS


##########################################  ETAPE 1  ##########################################
#OPTION SYMBOLGEN MPRINT MLOGIC

TEST_MIDCORP_PTF_IMS = PERIM_PTF_IMS.copy()

TEST_MIDCORP_PTF_IMS["COUNT_LIGNE"] = 1
TEST_MIDCORP_PTF_IMS["BUS_VAL_COVERAGE_PERIL"] = 0

for index, row in TEST_MIDCORP_PTF_IMS.iterrows():
    for i in range(1, 29):
        coli = f"LBGAR{i}"
        if row.get(coli) is not None:
            for j in range(1, 29):
                colj = f"MTPRPR{j}"
                if row.get(colj) is not None and row[colj] > 0:
                    TEST_MIDCORP_PTF_IMS.at[index, "BUS_VAL_COVERAGE_PERIL"] += 1
                    break
            if TEST_MIDCORP_PTF_IMS.at[index, "BUS_VAL_COVERAGE_PERIL"] > 0:
                break


##########################################  ETAPE 2  ##########################################

RESULTS_MIDCORP_PTF_IMS = pd.DataFrame({"SUM_COUNT_LIGNE" : [TEST_MIDCORP_PTF_IMS["COUNT_LIGNE"].sum()]})
RESULTS_MIDCORP_PTF_IMS["SUM_BUS_VAL_COVERAGE_PERIL"] = TEST_MIDCORP_PTF_IMS["BUS_VAL_COVERAGE_PERIL"].sum()


# In[18]:


#RESULTS_MIDCORP_PTF_IMS


# In[20]:


from openpyxl import load_workbook
import pandas as pd

chemin = './test/DataFitness2024_DATAFY_P4D_Didier 1 (version 1).xlsx'

workbook = load_workbook(chemin)


NouvelleFeuille = f"MIDCORP_PTF_IMS_{annee}{mois}"


nouvelle_feuille = workbook.create_sheet(title= NouvelleFeuille)

nouvelle_feuille.append(list(RESULTS_MIDCORP_PTF_IMS.columns))

for index, row in RESULTS_MIDCORP_PTF_IMS.iterrows():
    nouvelle_feuille.append(list(row))

workbook.save(chemin)

print(f"Le DataFrame a été ajouté dans {chemin} sous la feuille '{NouvelleFeuille}'")


workbook.close()


# In[21]:


######################################################################################################################
####################################/****** DTM MidCorp : Table MVT Lapses ******/####################################
######################################################################################################################

#####  Policy   ######
PERIM_MVT_MC = MVT_MC.copy()

VarMVLapses = ["NOPOL", "NOINT", "CDPOLE", "CDPROD", "DTCREPOL", "DTRESILP", "CDSITP"]


##########################################  ETAPE 1  ##########################################

TEST_MIDCORP_LAPSES_MVT = PERIM_MVT_MC[VarMVLapses]

#################    Nombre de ligne au total   ##################################
TEST_MIDCORP_LAPSES_MVT["COUNT_LIGNES"] = 1

#################   NOPOL  ##################################
COMPLETENESS(TEST_MIDCORP_LAPSES_MVT, "NOPOL")
TECH_VAL_FORMAT_VAR(TEST_MIDCORP_LAPSES_MVT, "NOPOL")
TECH_VALIDITY_LENGTH(TEST_MIDCORP_LAPSES_MVT, "NOPOL")

#################   NOINT  ##################################
COMPLETENESS(TEST_MIDCORP_LAPSES_MVT, "NOINT")
TECH_VAL_FORMAT_VAR(TEST_MIDCORP_LAPSES_MVT, "NOINT")
TECH_VALIDITY_LENGTH(TEST_MIDCORP_LAPSES_MVT, "NOINT")

#################   CDPOLE  ##################################
COMPLETENESS(TEST_MIDCORP_LAPSES_MVT, "CDPOLE")
BUSINESS_VALIDY (TEST_MIDCORP_LAPSES_MVT, "CDPOLE")

#################   CDPROD  ##################################
COMPLETENESS(TEST_MIDCORP_LAPSES_MVT, "CDPROD")
#BUSINESS_VALIDY (TEST_MIDCORP_LAPSES_MVT, "CDPROD")

#################   DTCREPOL  ##################################
COMPLETENESS(TEST_MIDCORP_LAPSES_MVT, "DTCREPOL")


#BUSINESS_VALIDY (TEST_MIDCORP_LAPSES_MVT, "DTCREPOL")  EN ATTENTE D'UN ECLAIRCISSEMENT SUR LE FAIT QUE LA COLONE "DTEFSITT"  N'EXISTE PAS DANS LA TABLE MVT_MC
TEST_MIDCORP_LAPSES_MVT["BUS_VAL_DTCREPOL"] = 0  

#################   DTRESILP  ##################################
# Spécifique car on souhaite regarder uniquemment les enregistrements qui ont un statut de police à Cancelled / RESILIé 


#COUNT_POLICY_CANCELLED
TEST_MIDCORP_LAPSES_MVT["COUNT_POLICY_CANCELLED"] = np.where(TEST_MIDCORP_LAPSES_MVT["CDSITP"] == '3', 1, 0)

TEST_MIDCORP_LAPSES_MVT["COMP_DTRESILP"] = 0
TEST_MIDCORP_LAPSES_MVT.loc[TEST_MIDCORP_LAPSES_MVT["COUNT_POLICY_CANCELLED"] == 1, "COMP_DTRESILP"] = TEST_MIDCORP_LAPSES_MVT.loc[TEST_MIDCORP_LAPSES_MVT["COUNT_POLICY_CANCELLED"] == 1, "DTRESILP"].apply(lambda t: int(not (pd.isna(t) | (t == ''))))

#BUS_VAL_DTRESILP
# TEST_MIDCORP_LAPSES_MVT["BUS_VAL_DTRESILP"] = 0

# TEST_MIDCORP_LAPSES_MVT["A"] = np.where(((TEST_MIDCORP_LAPSES_MVT["COMP_DTRESILP"] == 1) & (TEST_MIDCORP_LAPSES_MVT["COMP_DTEFSITT"] == 1)), 1, 0)
# df = TEST_MIDCORP_LAPSES_MVT[TEST_MIDCORP_LAPSES_MVT["A"] == 1]
# df.loc[df["COUNT_POLICY_CANCELLED"] == 1, "BUS_VAL_DTRESILP"] = df[df["COUNT_POLICY_CANCELLED"] == 1].apply(lambda t: pd.to_datetime(t["DTRESILP"]) >= pd.to_datetime(t["DTEFSITT"]), axis = 1)
# TEST_MIDCORP_LAPSES_MVT["BUS_VAL_DTRESILP"]  =  df["BUS_VAL_DTRESILP"].astype(int)  # je le transforme en int ici d'abord comme ca il ne reste que les NaN à changer

# TEST_MIDCORP_LAPSES_MVT["BUS_VAL_DTRESILP"].fillna(0, inplace=True) # je le transforme ici si non il va me transforme aussi les Fals en 0
# TEST_MIDCORP_LAPSES_MVT["BUS_VAL_DTRESILP"] = TEST_MIDCORP_LAPSES_MVT["BUS_VAL_DTRESILP"].astype(int)

TEST_MIDCORP_LAPSES_MVT["BUS_VAL_DTRESILP"] = 0     ##   EN ATTENTE D'UN ECLAIRCISSEMENT SUR LE FAIT QUE LA COLONE "DTEFSITT"  N4EXISTEPAS DANS LA TABLE MVT_MC

#################   CDSITP  ##################################
COMPLETENESS(TEST_MIDCORP_LAPSES_MVT, "CDSITP")
BUSINESS_VALIDY (TEST_MIDCORP_LAPSES_MVT, "CDSITP")


TEST_MIDCORP_LAPSES_MVT = TEST_MIDCORP_LAPSES_MVT.drop(columns = VarMVLapses)
#TEST_MIDCORP_LAPSES_MVT


#################    ETAPE 2  ##################################

############## Nombre de ligne au total ###########################
RESULTS_MIDCORP_LAPSES_MVT = pd.DataFrame({"SUM_COUNT_LIGNES" : [(TEST_MIDCORP_LAPSES_MVT["COUNT_LIGNES"]).sum()]})   # Il faut obligatoirement faire ceci pour avoir la premiere ligne et le reste suivra


############## NOPOL ###########################
RESULTS_MIDCORP_LAPSES_MVT[f'{REFMIDCORP["NOPOL"]}_COMP'] = TEST_MIDCORP_LAPSES_MVT["COMP_NOPOL"].sum()	
RESULTS_MIDCORP_LAPSES_MVT[f'{REFMIDCORP["NOPOL"]}_FORMAT'] = TEST_MIDCORP_LAPSES_MVT["FORMAT_NOPOL"].sum()	
RESULTS_MIDCORP_LAPSES_MVT[f'{REFMIDCORP["NOPOL"]}_LENGTH'] = TEST_MIDCORP_LAPSES_MVT["LENGTH_NOPOL"].sum()	


############## NOINT ###########################
RESULTS_MIDCORP_LAPSES_MVT[f'{REFMIDCORP["NOINT"]}_COMP'] = TEST_MIDCORP_LAPSES_MVT["COMP_NOINT"].sum()	
RESULTS_MIDCORP_LAPSES_MVT[f'{REFMIDCORP["NOINT"]}_FORMAT'] = TEST_MIDCORP_LAPSES_MVT["FORMAT_NOINT"].sum()	
RESULTS_MIDCORP_LAPSES_MVT[f'{REFMIDCORP["NOINT"]}_LENGTH'] = TEST_MIDCORP_LAPSES_MVT["LENGTH_NOINT"].sum()	


############## CDPOLE ###########################
RESULTS_MIDCORP_LAPSES_MVT[f'{REFMIDCORP["CDPOLE"]}_COMP'] = TEST_MIDCORP_LAPSES_MVT["COMP_CDPOLE"].sum()	
RESULTS_MIDCORP_LAPSES_MVT[f'{REFMIDCORP["CDPOLE"]}_BUS_VAL'] = TEST_MIDCORP_LAPSES_MVT["BUS_VAL_CDPOLE"].sum()	


############## CDPROD ###########################
RESULTS_MIDCORP_LAPSES_MVT[f'{REFMIDCORP["CDPROD"]}_COMP'] = TEST_MIDCORP_LAPSES_MVT["COMP_CDPROD"].sum()	

# BUS_VAL_CDPROD spécifique
CDPROD_distinc = SEGPROD["CPROD"].unique()
filtre = PERIM_MVT_MC[PERIM_MVT_MC['CDPROD'].isin(CDPROD_distinc)]

RESULTS_MIDCORP_LAPSES_MVT[f'{REFMIDCORP["CDPROD"]}_BUS_VAL'] = len(filtre)

############## DTCREPOL ###########################
RESULTS_MIDCORP_LAPSES_MVT[f'{REFMIDCORP["DTCREPOL"]}_COMP'] = TEST_MIDCORP_LAPSES_MVT["COMP_DTCREPOL"].sum()	
RESULTS_MIDCORP_LAPSES_MVT[f'{REFMIDCORP["DTCREPOL"]}_BUS_VAL'] = TEST_MIDCORP_LAPSES_MVT["BUS_VAL_DTCREPOL"].sum()


############## DTRESILP ###########################
RESULTS_MIDCORP_LAPSES_MVT[f'{REFMIDCORP["COUNT_POLICY_CANCELLED"]}_COMP'] = TEST_MIDCORP_LAPSES_MVT["COUNT_POLICY_CANCELLED"].sum()	
RESULTS_MIDCORP_LAPSES_MVT[f'{REFMIDCORP["DTRESILP"]}_COMP'] = TEST_MIDCORP_LAPSES_MVT["COMP_DTRESILP"].sum()	
RESULTS_MIDCORP_LAPSES_MVT[f'{REFMIDCORP["DTRESILP"]}_BUS_VAL'] = TEST_MIDCORP_LAPSES_MVT["BUS_VAL_DTRESILP"].sum()	


############## CDSITP ###########################
RESULTS_MIDCORP_LAPSES_MVT[f'{REFMIDCORP["CDSITP"]}_COMP'] = TEST_MIDCORP_LAPSES_MVT["COMP_CDSITP"].sum()	
RESULTS_MIDCORP_LAPSES_MVT[f'{REFMIDCORP["CDSITP"]}_BUS_VAL'] = TEST_MIDCORP_LAPSES_MVT["BUS_VAL_CDSITP"].sum()	

#RESULTS_MIDCORP_LAPSES_MVT


# In[22]:


from openpyxl import load_workbook
import pandas as pd

chemin = './test/DataFitness2024_DATAFY_P4D_Didier 1 (version 1).xlsx'

workbook = load_workbook(chemin)


NouvelleFeuille = f"MIDCORP_LAPSES_MVT_{annee}{mois}"


nouvelle_feuille = workbook.create_sheet(title= NouvelleFeuille)

nouvelle_feuille.append(list(RESULTS_MIDCORP_LAPSES_MVT.columns))

for index, row in RESULTS_MIDCORP_LAPSES_MVT.iterrows():
    nouvelle_feuille.append(list(row))

workbook.save(chemin)

print(f"Le DataFrame a été ajouté dans {chemin} sous la feuille '{NouvelleFeuille}'")


workbook.close()


# In[24]:


#################   Spécifique au CDPROD  ##################################

CDPROD_distinc = SEGPROD["CPROD"].unique()
filtre = PERIM_MVT_MC[PERIM_MVT_MC['CDPROD'].isin(CDPROD_distinc)]

RESULT_BUS_VAL_CDPROD = pd.DataFrame({f'{REFMIDCORP["CDPROD"]}_COMP' : [len(filtre)]})

#RESULT_BUS_VAL_CDPROD   #cette DF a été rajouté sur la plus grande DF nommé "RESULTS_MIDCORP_LAPSES_MVT"


# In[25]:


###############################################              FIN POUR DAMOCLES               ###################################

