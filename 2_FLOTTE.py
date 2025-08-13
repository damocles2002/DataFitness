#!/usr/bin/env python
# coding: utf-8

# In[1]:


###############################################################################################################################

# Lancement des tests

# Le programme fonctionne en 2 étapes :
# 1) Il va lire la table et va créer des variables qui vont prendre la valeur 1 si le test est OK, 0 sinon.
# 2) Ensuite, il va sommer chaque colonne et la table finale va donc contenir le nombre d'enregistrements OK pour un test donné.

###############################################################################################################################

###############################################################################################################################
####################################### DTM MidCorp : Table PTF POLICY ########################################
###############################################################################################################################


get_ipython().run_line_magic('run', 'LIEN_FONCT.ipynb')


# In[2]:


########################################   POLICY    ########################################

NEW_MVT_FL = MVT_FL.copy()

varMVTPOL = ["NOPOL",
	"NOCLT", 
	"NOINT",
	"DIRCOM",
    "CDPOLE",
	"TYPE",
    "SITU_POL",
    "dtcrepol",
    "dtresilp",
    "CMARCH",
    "CSEG",
    "CSSSEG",
    "POSRISQ"
    ]


TEST_FLOTTE_POLICY_MVT = NEW_MVT_FL[varMVTPOL]

######################################################   ETAPE 1  ################################


########################################   Nombre de ligne au total    ########################################

TEST_FLOTTE_POLICY_MVT["COUNT_LIGNES"] = 1


########################################  NOPOL    ########################################

COMPLETENESS(TEST_FLOTTE_POLICY_MVT, "NOPOL")
TECH_VAL_FORMAT_VAR(TEST_FLOTTE_POLICY_MVT, "NOPOL")
TECH_VALIDITY_LENGTH(TEST_FLOTTE_POLICY_MVT, "NOPOL")


########################################   NOCLT   ########################################

COMPLETENESS(TEST_FLOTTE_POLICY_MVT, "NOCLT")
TECH_VAL_FORMAT_VAR(TEST_FLOTTE_POLICY_MVT, "NOCLT")
TECH_VALIDITY_LENGTH(TEST_FLOTTE_POLICY_MVT, "NOCLT")


########################################   SITU_POL   ########################################

COMPLETENESS(TEST_FLOTTE_POLICY_MVT, "SITU_POL")
BUSINESS_VALIDY (TEST_FLOTTE_POLICY_MVT, "SITU_POL")


########################################   DTCREPOL   ########################################
#COMPLETNESS  JE LE FAIS LITTERALEMET CAR SUR LE COLONNE DE LA TABLE SAS LE NOM N'EST PAS ECRIT EN MAJUSCULE, MAIS IL ME FAUT COMPARER
TEST_FLOTTE_POLICY_MVT["COMP_DTCREPOL"] = np.where(pd.isna(TEST_FLOTTE_POLICY_MVT["dtcrepol"]) | (TEST_FLOTTE_POLICY_MVT["dtcrepol"] == ""), 0, 1)

# BUSINESS_VALIDY (TEST_FLOTTE_POLICY_MVT, "dtcrepol")    la colonne DTEFSITT n'existe pas dans la table MVT_FL


########################################   DTRESILP   ########################################

#COUNT_POLICY_CANCELLED
TEST_FLOTTE_POLICY_MVT["COUNT_POLICY_CANCELLED"] = np.where(TEST_FLOTTE_POLICY_MVT["SITU_POL"] == "RESILIEE", 1, 0)  


#COMPLETENESS
TEST_FLOTTE_POLICY_MVT["COMP_DTRESILP"] = 0  
TEST_FLOTTE_POLICY_MVT.loc[TEST_FLOTTE_POLICY_MVT["COUNT_POLICY_CANCELLED"] == 1, "COMP_DTRESILP"] = TEST_FLOTTE_POLICY_MVT.loc[TEST_FLOTTE_POLICY_MVT["COUNT_POLICY_CANCELLED"] == 1, "dtresilp"].apply(lambda t: int(not (pd.isna(t) or (t == ''))))


#BUS_VAL_DTRESILP
TEST_FLOTTE_POLICY_MVT["BUS_VAL_DTRESILP"] = 0   # la colonne DTEFSITT n'existe pas dans la table MVT_FL

# TEST_FLOTTE_POLICY_MVT["A"] = np.where(((TEST_FLOTTE_POLICY_MVT["COMP_DTRESILP"] == 1) & (TEST_FLOTTE_POLICY_MVT["COMP_DTEFSITT"] == 1)), 1, 0)
# df = TEST_FLOTTE_POLICY_MVT[TEST_FLOTTE_POLICY_MVT["A"] == 1]
# df.loc[df["COUNT_POLICY_CANCELLED"] == 1, "BUS_VAL_DTRESILP"] = df[df["COUNT_POLICY_CANCELLED"] == 1].apply(lambda t: pd.to_datetime(t["dtresilp"]) >= pd.to_datetime(t["DTEFSITT"]), axis = 1)
# TEST_FLOTTE_POLICY_MVT["BUS_VAL_DTRESILP"]  =  df["BUS_VAL_DTRESILP"].astype(int)  # je le transforme en int ici d'abord comme ca il ne reste que les NaN à changer

# TEST_FLOTTE_POLICY_MVT["BUS_VAL_DTRESILP"].fillna(0, inplace=True) # je le transforme ici si non il va me transforme aussi les Fals en 0
# TEST_FLOTTE_POLICY_MVT["BUS_VAL_DTRESILP"] = TEST_FLOTTE_POLICY_MVT["BUS_VAL_DTRESILP"].astype(int)


########################################   NOINT   ########################################

COMPLETENESS(TEST_FLOTTE_POLICY_MVT, "NOINT")
TECH_VAL_FORMAT_VAR(TEST_FLOTTE_POLICY_MVT, "NOINT")


#TECH_VAL_LENGTH_NOINT
TEST_FLOTTE_POLICY_MVT["LENGTH_NOINT"] = np.where((TEST_FLOTTE_POLICY_MVT["COMP_NOINT"] == 1) & (TEST_FLOTTE_POLICY_MVT["NOINT"].apply(lambda t: len(t)) == 5), 1, 0)


########################################   DIRCOM   ########################################

COMPLETENESS(TEST_FLOTTE_POLICY_MVT, "DIRCOM")

########################################   CDPOLE   ########################################

COMPLETENESS(TEST_FLOTTE_POLICY_MVT, "CDPOLE")
BUSINESS_VALIDY (TEST_FLOTTE_POLICY_MVT, "CDPOLE")


########################################   TYPE   ########################################

COMPLETENESS(TEST_FLOTTE_POLICY_MVT, "TYPE")

########################################   (CMARCH, CSEG, CSSSEG) = CCC   ########################################

COMPLETENESS(TEST_FLOTTE_POLICY_MVT, "CMARCH")
COMPLETENESS(TEST_FLOTTE_POLICY_MVT, "CSEG")
COMPLETENESS(TEST_FLOTTE_POLICY_MVT, "CSSSEG")

TEST_FLOTTE_POLICY_MVT["COMP_CCC"] = np.where((TEST_FLOTTE_POLICY_MVT["COMP_CMARCH"] == 1)  & (TEST_FLOTTE_POLICY_MVT["COMP_CSEG"] == 1
) & (TEST_FLOTTE_POLICY_MVT["COMP_CSSSEG"] == 1
), 1, 0)

BUSINESS_VALIDY(TEST_FLOTTE_POLICY_MVT, "CMARCH")
BUSINESS_VALIDY(TEST_FLOTTE_POLICY_MVT, "CSEG")
BUSINESS_VALIDY(TEST_FLOTTE_POLICY_MVT, "CSSSEG")

TEST_FLOTTE_POLICY_MVT["BUS_VAL_CCC"] = np.where((TEST_FLOTTE_POLICY_MVT["BUS_VAL_CMARCH"] == 1)  & (TEST_FLOTTE_POLICY_MVT["BUS_VAL_CSEG"] == 1
) & (TEST_FLOTTE_POLICY_MVT["BUS_VAL_CSSSEG"] == 1
), 1, 0)

########################################   POSRISQ   ########################################
COMPLETENESS(TEST_FLOTTE_POLICY_MVT, "POSRISQ")
TEST_FLOTTE_POLICY_MVT["BUS_VAL_POSRISQ"] = np.where(TEST_FLOTTE_POLICY_MVT["POSRISQ"].isin(LISTE_CODE_POSTAUX["CODE_POSTAL"]), 1, 0)



# In[3]:


TEST_FLOTTE_POLICY_MVT = TEST_FLOTTE_POLICY_MVT.drop(columns = varMVTPOL)
#TEST_FLOTTE_POLICY_MVT


# In[5]:


######################################################   ETAPE 2  ################################


############## Nombre de ligne au total ###########################
RESULTS_FLOTTE_POLICY_MVT = pd.DataFrame({"SUM_COUNT_LIGNES" : [(TEST_FLOTTE_POLICY_MVT["COUNT_LIGNES"]).sum()]})   # Il faut obligatoirement faire ceci pour avoir la premiere ligne et le reste suivra


####################### NOPOL ###########################
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["NOPOL"]}_COMP'] = TEST_FLOTTE_POLICY_MVT["COMP_NOPOL"].sum()	
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["NOPOL"]}_FORMAT'] = TEST_FLOTTE_POLICY_MVT["FORMAT_NOPOL"].sum()	
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["NOPOL"]}_LENGTH'] = TEST_FLOTTE_POLICY_MVT["LENGTH_NOPOL"].sum()	


####################### NOCLT ###########################
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["NOCLT"]}_COMP'] = TEST_FLOTTE_POLICY_MVT["COMP_NOCLT"].sum()	
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["NOCLT"]}_FORMAT'] = TEST_FLOTTE_POLICY_MVT["FORMAT_NOCLT"].sum()	
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["NOCLT"]}_LENGTH'] = TEST_FLOTTE_POLICY_MVT["LENGTH_NOCLT"].sum()	


####################### SITU_POL ###########################
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["SITU_POL"]}_COMP'] = TEST_FLOTTE_POLICY_MVT["COMP_SITU_POL"].sum()	
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["SITU_POL"]}_BUS_VAL'] = TEST_FLOTTE_POLICY_MVT["BUS_VAL_SITU_POL"].sum()


####################### DTCREPOL ###########################
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["DTCREPOL"]}_COMP'] = TEST_FLOTTE_POLICY_MVT["COMP_DTCREPOL"].sum()	


RESULTS_FLOTTE_POLICY_MVT["SUM_COUNT_POLICY_CANCELLED"] = TEST_FLOTTE_POLICY_MVT["COUNT_POLICY_CANCELLED"].sum()

RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["DTRESILP"]}_COMP'] = 0
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["DTRESILP"]}_BUS_VAL'] = 0


####################### DTRESILP ###########################
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["DTRESILP"]}_COMP'] = TEST_FLOTTE_POLICY_MVT["COMP_DTRESILP"].sum()	
# RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["DTRESILP"]}_BUS_VAL'] = TEST_FLOTTE_POLICY_MVT["BUS_VAL_DTRESILP"].sum()  # la colonne DTEFSITT n'existe pas dans la table MVT_FL


####################### NOINT ###########################
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["NOINT"]}_COMP'] = TEST_FLOTTE_POLICY_MVT["COMP_NOINT"].sum()	
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["NOINT"]}_FORMAT'] = TEST_FLOTTE_POLICY_MVT["FORMAT_NOINT"].sum()	
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["NOINT"]}_LENGTH'] = TEST_FLOTTE_POLICY_MVT["LENGTH_NOINT"].sum()	


####################### DIRCOM ###########################
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["DIRCOM"]}_COMP'] = TEST_FLOTTE_POLICY_MVT["COMP_DIRCOM"].sum()	

####################### CDPOLE ###########################
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["DIRCOM"]}_COMP'] = TEST_FLOTTE_POLICY_MVT["COMP_CDPOLE"].sum()	
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["DIRCOM"]}_BUS_VAL'] = TEST_FLOTTE_POLICY_MVT["BUS_VAL_CDPOLE"].sum()	


####################### TYPE ###########################
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["TYPE"]}_COMP'] = TEST_FLOTTE_POLICY_MVT["COMP_TYPE"].sum()

####################### (CMARCH, CSEG, CSSSEG) = CCC ###########################
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["CCC"]}_COMP'] = TEST_FLOTTE_POLICY_MVT["COMP_CCC"].sum()
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["CCC"]}_BUS_VAL'] = TEST_FLOTTE_POLICY_MVT["BUS_VAL_CCC"].sum()

########################################   POSRISQ   ########################################
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["POSRISQ"]}_COMP'] = TEST_FLOTTE_POLICY_MVT["COMP_POSRISQ"].sum()
RESULTS_FLOTTE_POLICY_MVT[f'{REFFLOTTE["POSRISQ"]}_BUS_VAL'] = TEST_FLOTTE_POLICY_MVT["BUS_VAL_POSRISQ"].sum()




# In[6]:


RESULTS_FLOTTE_POLICY_MVT


# In[8]:


from openpyxl import load_workbook
import pandas as pd


# In[9]:


chemin = './test/DataFitness2024_DATAFY_P4D_Didier 1 (version 1).xlsx'

workbook = load_workbook(chemin)


# In[10]:


NouvelleFeuille = f"Flotte_MVT_{annee}{mois}"


nouvelle_feuille = workbook.create_sheet(title= NouvelleFeuille)

nouvelle_feuille.append(list(RESULTS_FLOTTE_POLICY_MVT.columns))

for index, row in RESULTS_FLOTTE_POLICY_MVT.iterrows():
    nouvelle_feuille.append(list(row))

workbook.save(chemin)

print(f"Le DataFrame a été ajouté dans {chemin} sous la feuille '{NouvelleFeuille}'")


# In[11]:


workbook.close()


# In[12]:


liste_feuilles = workbook.sheetnames


# In[13]:


print(liste_feuilles)


# In[15]:


####################### Pour vérifier les garanties ###########################
"""
#trouve le lien de la table "GARANTIE FLOTTE" PUIS JE POURRAIS VERIFIER CE QUE JE FAIS 

TEST_FLOTTE_POLICY_GAR = GAR_FL.copy()

####################### Nombre de ligne au total ###########################
TEST_FLOTTE_POLICY_GAR["COUNT_LIGNES"] = 1

TEST_FLOTTE_POLICY_GAR["COUNT_GAR"] = 0
for index, row in TEST_FLOTTE_POLICY_GAR:
    L = TEST_FLOTTE_POLICY_GAR.columns
    for i in L:
        if any(row.get(i).notna() & row.get(i).startswith('GAR')):
            TEST_FLOTTE_POLICY_GAR[index, "COUNT_GAR"] = 1

"""

#startswith('GAR') verifier si le caracter commence par "GAR"


"""
RESULTS_FLOTTE_POLICY_GAR = pd.DataFrame({"TT_COUNT_LIGNES" : [TEST_FLOTTE_POLICY_GAR["COUNT_LIGNES"].sum()]})

####################### NOPOL ###########################
RESULTS_FLOTTE_POLICY_GAR["TT_COUNT_GAR"] = TEST_FLOTTE_POLICY_GAR["COUNT_GAR"].sum()
"""


# In[16]:


##############################################################              FIN POUR DAMOCLES               #######################################################################

